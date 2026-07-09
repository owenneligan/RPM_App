"""
CSI Lookup Tool generator.
Run this script to produce CSI_Tool.xlsx — a self-contained Excel workbook
for insurance call-centre staff to look up average Contents Sum Insured by
UK postcode area and bedroom count.
"""

import zipfile, shutil, os, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

# ── postcode centroids ────────────────────────────────────────────────────────
from postcode_coords import POSTCODE_COORDS

OUT = Path("/tmp/claude-0/-home-user-RPM-App/46ca97ba-bf60-527c-b9f7-3e13f6d92b04/scratchpad/CSI_Tool.xlsx")

# ── colour palette ────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"
GOLD   = "C9A84C"
LIGHT  = "F5F7FA"
WHITE  = "FFFFFF"
GREEN  = "2E7D32"
GREY   = "B0BEC5"
LGREY  = "ECEFF1"

def side(c="D0D0D0"): return Side(style="thin", color=c)
def border(c="D0D0D0"): return Border(left=side(c), right=side(c),
                                       top=side(c), bottom=side(c))

# =============================================================================
# SHEET 1 — DATA
# =============================================================================
def build_data_sheet(wb):
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = True
    ws.sheet_properties.tabColor = NAVY

    # ── Row 1: instruction banner ─────────────────────────────────────────
    note = ws.cell(1, 1,
        "PASTE YOUR DATA BELOW (row 3 onward).  "
        "Columns: Postcode | Bedrooms | CSI_Value.  "
        "After pasting new data go to  Data > Refresh All  (or Ctrl+Alt+F5).")
    note.font = Font(italic=True, color="555555", size=10)
    note.fill = PatternFill("solid", fgColor=LGREY)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 18

    # ── Row 2: header row ─────────────────────────────────────────────────
    headers = ["Postcode", "Bedrooms", "CSI_Value"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border(NAVY)
    ws.row_dimensions[2].height = 22

    # ── Rows 3+: sample data ──────────────────────────────────────────────
    samples = [
        ("SE3", 3, 125000), ("SE3", 4, 175000), ("SE3", 3, 110000),
        ("SW1", 5, 350000), ("SW1", 4, 280000), ("SW1", 5, 320000),
        ("E1",  2, 85000),  ("E1",  2, 92000),  ("E1",  3, 130000),
        ("N1",  3, 145000), ("N1",  4, 195000), ("M1",  2, 65000),
        ("M1",  3, 90000),  ("LS1", 2, 60000),  ("LS1", 3, 85000),
    ]
    for r, (pc, beds, csi) in enumerate(samples, 3):
        ws.cell(r, 1, pc.upper())
        ws.cell(r, 2, beds)
        c = ws.cell(r, 3, csi)
        c.number_format = '£#,##0'

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    # Named table starting at row 2 (tblData — Power Query reads this)
    last_row = len(samples) + 2
    tbl = Table(displayName="tblData",
                ref=f"A2:C{last_row}",
                tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",
                                              showRowStripes=True))
    ws.add_table(tbl)

    # Freeze below header row
    ws.freeze_panes = "A3"

    return ws


# =============================================================================
# SHEET 2 — POSTCODE COORDS (hidden reference)
# =============================================================================
def build_coords_sheet(wb):
    ws = wb.create_sheet("PostcodeCoords")
    ws.sheet_state = "hidden"

    headers = ["Postcode", "Lat", "Lon"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True)

    for row, (pc, (lat, lon)) in enumerate(sorted(POSTCODE_COORDS.items()), 2):
        ws.cell(row, 1, pc)
        ws.cell(row, 2, lat)
        ws.cell(row, 3, lon)

    last = len(POSTCODE_COORDS) + 1
    tbl = Table(displayName="tblPostcodeCoords", ref=f"A1:C{last}",
                tableStyleInfo=TableStyleInfo(name="TableStyleLight1"))
    ws.add_table(tbl)
    return ws


# =============================================================================
# SHEET 3 — PQ_SUMMARY (Power Query output lands here)
# =============================================================================
def build_pq_summary_sheet(wb):
    ws = wb.create_sheet("PQ_Summary")
    ws.sheet_state = "hidden"

    headers = ["Postcode", "Bedrooms", "AvgCSI", "Count"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)

    # Placeholder row so Excel recognises the table shape
    placeholders = [("SE3", 3, 117500, 2), ("SW1", 5, 335000, 2),
                    ("E1",  2, 88500,  2), ("N1",  3, 145000, 1),
                    ("M1",  2, 65000,  1), ("LS1", 2, 60000,  1)]
    for r, row in enumerate(placeholders, 2):
        for col, val in enumerate(row, 1):
            ws.cell(r, col, val)

    last = len(placeholders) + 1
    tbl = Table(displayName="tblSummary", ref=f"A1:D{last}",
                tableStyleInfo=TableStyleInfo(name="TableStyleLight1"))
    ws.add_table(tbl)
    return ws


# =============================================================================
# SHEET 4 — PQ_POSTAGG (postcode-level CSI per bedroom)
# =============================================================================
def build_pq_agg_sheet(wb):
    ws = wb.create_sheet("PQ_PostcodeAgg")
    ws.sheet_state = "hidden"

    headers = ["Postcode", "AvgCSI", "AvgBeds", "CSIPerBedroom"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h).font = Font(bold=True)

    placeholders = [
        ("SE3",  117500, 3.0, 39167),
        ("SW1",  307500, 4.5, 68333),
        ("E1",   102333, 2.3, 44493),
        ("N1",   145000, 3.0, 48333),
        ("M1",    65000, 2.0, 32500),
        ("LS1",   60000, 2.0, 30000),
    ]
    for r, row in enumerate(placeholders, 2):
        for col, val in enumerate(row, 1):
            ws.cell(r, col, val)

    last = len(placeholders) + 1
    tbl = Table(displayName="tblPostcodeAgg", ref=f"A1:D{last}",
                tableStyleInfo=TableStyleInfo(name="TableStyleLight1"))
    ws.add_table(tbl)
    return ws


# =============================================================================
# SHEET 5 — TOOL (the call-handler UI)
# =============================================================================
def build_tool_sheet(wb):
    ws = wb.create_sheet("Tool")
    ws.sheet_properties.tabColor = GOLD

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 3

    # Row heights
    for r in range(1, 30):
        ws.row_dimensions[r].height = 20
    ws.row_dimensions[3].height = 32   # title
    ws.row_dimensions[8].height = 26   # input labels
    ws.row_dimensions[9].height = 32   # input fields
    ws.row_dimensions[11].height = 32  # bedrooms input
    ws.row_dimensions[14].height = 50  # result
    ws.row_dimensions[17].height = 24  # sample size
    ws.row_dimensions[19].height = 22  # note

    # ── HEADER BANNER ────────────────────────────────────────────────────────
    ws.merge_cells("B3:C3")
    title = ws.cell(3, 2, "Contents Sum Insured Lookup Tool")
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # ── INPUT LABELS ─────────────────────────────────────────────────────────
    def label(row, col, text):
        c = ws.cell(row, col, text)
        c.font = Font(bold=True, color=NAVY, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")

    label(8, 2, "Postcode Area (e.g. SE3, SW1A)")
    label(10, 2, "Number of Bedrooms")

    # ── INPUT CELLS ──────────────────────────────────────────────────────────
    def input_cell(row, col):
        c = ws.cell(row, col)
        c.fill = PatternFill("solid", fgColor="FFF9E6")
        c.font = Font(size=13, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            left=Side(style="medium", color=GOLD),
            right=Side(style="medium", color=GOLD),
            top=Side(style="medium", color=GOLD),
            bottom=Side(style="medium", color=GOLD),
        )
        return c

    pc_cell  = input_cell(9, 2)   # B9
    bed_cell = input_cell(11, 2)  # B11

    # Default placeholder values
    pc_cell.value  = "SE3"
    bed_cell.value = 3

    # ── RESULT HEADER ────────────────────────────────────────────────────────
    ws.merge_cells("B13:C13")
    rh = ws.cell(13, 2, "Suggested Contents Sum Insured")
    rh.font = Font(bold=True, size=11, color="555555")
    rh.alignment = Alignment(horizontal="center")

    # ── MAIN RESULT CELL (B14:C14) ───────────────────────────────────────────
    ws.merge_cells("B14:C14")
    result = ws.cell(14, 2)
    result.fill = PatternFill("solid", fgColor=NAVY)
    result.font = Font(bold=True, size=24, color=GOLD)
    result.alignment = Alignment(horizontal="center", vertical="center")
    result.border = Border(
        left=Side(style="medium", color=GOLD),
        right=Side(style="medium", color=GOLD),
        top=Side(style="medium", color=GOLD),
        bottom=Side(style="medium", color=GOLD),
    )
    result.number_format = '£#,##0'

    # ── SAMPLE SIZE (B16) ────────────────────────────────────────────────────
    ws.merge_cells("B16:C16")
    sample = ws.cell(16, 2)
    sample.font = Font(italic=True, color="555555", size=10)
    sample.alignment = Alignment(horizontal="center")

    # ── NOTE (B18) ───────────────────────────────────────────────────────────
    ws.merge_cells("B18:C18")
    note = ws.cell(18, 2)
    note.font = Font(italic=True, color="888888", size=9)
    note.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── DIVIDER LINE ─────────────────────────────────────────────────────────
    for col in [2, 3]:
        ws.cell(12, col).border = Border(
            bottom=Side(style="thin", color=GOLD))

    # ── FOOTER ───────────────────────────────────────────────────────────────
    ws.merge_cells("B20:C20")
    footer = ws.cell(20, 2,
        "For internal use only. Data refreshed from the Data sheet.")
    footer.font = Font(italic=True, size=8, color=GREY)
    footer.alignment = Alignment(horizontal="center")

    # ── FORMULAS ─────────────────────────────────────────────────────────────
    # Normalised postcode  (helper — stored in an unused cell)
    ws["E1"] = '=UPPER(TRIM(B9))'   # normalised postcode
    ws["E1"].font = Font(color=WHITE)  # invisible

    # The three-tier lookup — see comment block below for logic
    #
    # TIER 1: exact match in tblSummary (postcode + bedrooms)
    # TIER 2: postcode exists → extrapolate via tblPostcodeAgg CSIPerBedroom
    # TIER 3: geographic fallback → find nearest postcode in tblSummary
    #         using Euclidean distance on lat/lon, then re-apply tiers 1/2
    #
    # All wrapped in LET() for readability and to avoid recomputing.

    result.value = (
        '=LET('
        'pc,UPPER(TRIM(B9)),'
        'beds,B11,'

        # ── tier-1 exact match ──────────────────────────────────────────
        'exact,IFERROR(AVERAGEIFS(tblSummary[AvgCSI],'
                                 'tblSummary[Postcode],pc,'
                                 'tblSummary[Bedrooms],beds),0),'

        # ── tier-2 postcode exists, extrapolate ──────────────────────────
        'pc_rate,IFERROR(XLOOKUP(pc,tblPostcodeAgg[Postcode],'
                                 'tblPostcodeAgg[CSIPerBedroom],0),0),'
        'tier2,pc_rate*beds,'

        # ── tier-3 geographic fallback ───────────────────────────────────
        # get lat/lon of typed postcode from full coords table
        'tlat,IFERROR(XLOOKUP(pc,tblPostcodeCoords[Postcode],'
                              'tblPostcodeCoords[Lat],NA()),NA()),'
        'tlon,IFERROR(XLOOKUP(pc,tblPostcodeCoords[Postcode],'
                              'tblPostcodeCoords[Lon],NA()),NA()),'

        # unique postcodes that ARE in our summary data
        'avail,UNIQUE(tblSummary[Postcode]),'

        # lat/lon for each available postcode
        'alat,XLOOKUP(avail,tblPostcodeCoords[Postcode],'
                     'tblPostcodeCoords[Lat],0),'
        'alon,XLOOKUP(avail,tblPostcodeCoords[Postcode],'
                     'tblPostcodeCoords[Lon],0),'

        # squared Euclidean distance (no need for sqrt for min-finding)
        'dists,(alat-tlat)^2+(alon-tlon)^2,'
        'nearest,INDEX(avail,MATCH(MIN(dists),dists,0)),'

        # re-run tiers 1 & 2 with the nearest postcode
        'near_exact,IFERROR(AVERAGEIFS(tblSummary[AvgCSI],'
                                       'tblSummary[Postcode],nearest,'
                                       'tblSummary[Bedrooms],beds),0),'
        'near_rate,IFERROR(XLOOKUP(nearest,tblPostcodeAgg[Postcode],'
                                   'tblPostcodeAgg[CSIPerBedroom],0),0),'
        'near_tier,IF(near_exact>0,near_exact,near_rate*beds),'

        # ── resolve which tier to show ───────────────────────────────────
        'IF(exact>0,exact,IF(tier2>0,tier2,near_tier))'
        ')'
    )

    # Sample count formula
    sample.value = (
        '=LET('
        'pc,UPPER(TRIM(B9)),'
        'beds,B11,'
        'cnt,IFERROR(XLOOKUP(1,(tblSummary[Postcode]=pc)*(tblSummary[Bedrooms]=beds),'
                            'tblSummary[Count],0),0),'
        'IF(cnt>0,'
        '"Based on "&TEXT(cnt,"#,##0")&" similar properties",'
        '"Estimated — no exact match found for this postcode/bedroom combination")'
        ')'
    )

    # Explanatory note formula
    note.value = (
        '=LET('
        'pc,UPPER(TRIM(B9)),'
        'beds,B11,'
        'in_summary,COUNTIF(tblSummary[Postcode],pc)>0,'
        'exact_match,IFERROR(AVERAGEIFS(tblSummary[AvgCSI],'
                                        'tblSummary[Postcode],pc,'
                                        'tblSummary[Bedrooms],beds),0)>0,'
        'tlat,IFERROR(XLOOKUP(pc,tblPostcodeCoords[Postcode],'
                              'tblPostcodeCoords[Lat],NA()),NA()),'
        'tlon,IFERROR(XLOOKUP(pc,tblPostcodeCoords[Postcode],'
                              'tblPostcodeCoords[Lon],NA()),NA()),'
        'avail,UNIQUE(tblSummary[Postcode]),'
        'alat,XLOOKUP(avail,tblPostcodeCoords[Postcode],'
                     'tblPostcodeCoords[Lat],0),'
        'alon,XLOOKUP(avail,tblPostcodeCoords[Postcode],'
                     'tblPostcodeCoords[Lon],0),'
        'dists,(alat-tlat)^2+(alon-tlon)^2,'
        'nearest,IFERROR(INDEX(avail,MATCH(MIN(dists),dists,0)),""),'
        'IF(exact_match,"✓ Exact match",'
        'IF(in_summary,"~ Estimated by extrapolating postcode average",'
        'IF(nearest<>"",'
        '"⚠ Postcode not in data — using nearest area: "&nearest,'
        '"⚠ No data available")))'
        ')'
    )

    # Hide helper column E
    ws.column_dimensions["E"].hidden = True

    return ws


# =============================================================================
# EMBED POWER QUERY (inject XML into the xlsx ZIP)
# =============================================================================
PQ_SUMMARY_M = """\
let
    Source = Excel.CurrentWorkbook(){[Name="tblData"]}[Content],
    ChangedTypes = Table.TransformColumnTypes(Source, {
        {"Postcode",  type text},
        {"Bedrooms",  Int64.Type},
        {"CSI_Value", type number}
    }),
    Cleaned = Table.TransformColumns(ChangedTypes, {
        {"Postcode", each Text.Upper(Text.Trim(_)), type text}
    }),
    Grouped = Table.Group(Cleaned, {"Postcode", "Bedrooms"}, {
        {"AvgCSI", each List.Average([CSI_Value]), type number},
        {"Count",  each Table.RowCount(_),         Int64.Type}
    })
in
    Grouped"""

PQ_AGG_M = """\
let
    Source = Excel.CurrentWorkbook(){[Name="tblData"]}[Content],
    ChangedTypes = Table.TransformColumnTypes(Source, {
        {"Postcode",  type text},
        {"Bedrooms",  Int64.Type},
        {"CSI_Value", type number}
    }),
    Cleaned = Table.TransformColumns(ChangedTypes, {
        {"Postcode", each Text.Upper(Text.Trim(_)), type text}
    }),
    Grouped = Table.Group(Cleaned, {"Postcode"}, {
        {"AvgCSI",        each List.Average([CSI_Value]),                           type number},
        {"AvgBeds",       each List.Average([Bedrooms]),                            type number},
        {"CSIPerBedroom", each List.Average([CSI_Value]) / List.Average([Bedrooms]),type number}
    })
in
    Grouped"""


def inject_power_query(xlsx_path: Path):
    """
    Re-opens the xlsx as a ZIP and injects the two Power Query queries.
    Queries are stored in xl/customXml/item1.xml (Mashup/M code) and
    referenced via xl/connections.xml.
    """
    import zipfile, io

    # ── Build the Mashup XML that holds the M code ────────────────────────
    def escape(s):
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))

    mashup_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Workbook xmlns="http://schemas.microsoft.com/DataModel/Query">'
        '<Query Name="CSI_Summary">'
        f'<Formula>{escape(PQ_SUMMARY_M)}</Formula>'
        '</Query>'
        '<Query Name="CSI_PostcodeAgg">'
        f'<Formula>{escape(PQ_AGG_M)}</Formula>'
        '</Query>'
        '</Workbook>'
    )

    # ── Connections XML ───────────────────────────────────────────────────
    connections_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'

        # connection 1 → PQ_Summary sheet
        '<connection id="1" keepAlive="0" name="Query - CSI_Summary" '
        'description="Connection to the Power Query CSI_Summary" '
        'type="5" refreshedVersion="0" background="0" saveData="1" '
        'refreshOnLoad="1">'
        '<dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;'
        'Location=CSI_Summary" command="SELECT * FROM [CSI_Summary]"/>'
        '</connection>'

        # connection 2 → PQ_PostcodeAgg sheet
        '<connection id="2" keepAlive="0" name="Query - CSI_PostcodeAgg" '
        'description="Connection to the Power Query CSI_PostcodeAgg" '
        'type="5" refreshedVersion="0" background="0" saveData="1" '
        'refreshOnLoad="1">'
        '<dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;'
        'Location=CSI_PostcodeAgg" command="SELECT * FROM [CSI_PostcodeAgg]"/>'
        '</connection>'

        '</connections>'
    )

    # ── customXml relationship XML ────────────────────────────────────────
    custom_xml_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.microsoft.com/DataModel/Query" '
        'Target="item1.xml"/>'
        '</Relationships>'
    )

    # ── Patch the ZIP in memory ───────────────────────────────────────────
    buf = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:

        existing = set(zin.namelist())
        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == "xl/connections.xml":
                # Replace existing connections file
                zout.writestr(item, connections_xml.encode())
                continue

            if item.filename == "[Content_Types].xml":
                # Inject customXml content-type override
                ct = data.decode()
                inject = (
                    '<Override PartName="/xl/customXml/item1.xml" '
                    'ContentType="application/vnd.ms-excel.datamodelconnection+xml"/>'
                    '<Default Extension="rels" '
                    'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                )
                ct = ct.replace("</Types>", inject + "</Types>")
                zout.writestr(item, ct.encode())
                continue

            if item.filename == "xl/_rels/workbook.xml.rels":
                # Add relationship to customXml
                rels = data.decode()
                new_rel = (
                    '<Relationship Id="rIdCustomXml1" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml" '
                    'Target="../customXml/item1.xml"/>'
                    '<Relationship Id="rIdConnections" '
                    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections" '
                    'Target="connections.xml"/>'
                )
                rels = rels.replace("</Relationships>", new_rel + "</Relationships>")
                zout.writestr(item, rels.encode())
                continue

            zout.writestr(item, data)

        # Add missing files
        if "xl/connections.xml" not in existing:
            zout.writestr("xl/connections.xml", connections_xml.encode())

        zout.writestr("xl/customXml/item1.xml", mashup_xml.encode())
        zout.writestr("xl/customXml/_rels/item1.xml.rels", custom_xml_rels.encode())

    # Write back
    xlsx_path.write_bytes(buf.getvalue())
    print("  Power Query XML injected.")


# =============================================================================
# MAIN
# =============================================================================
def main():
    wb = Workbook()

    print("Building Data sheet …")
    build_data_sheet(wb)

    print("Building PostcodeCoords sheet …")
    build_coords_sheet(wb)

    print("Building PQ_Summary sheet …")
    build_pq_summary_sheet(wb)

    print("Building PQ_PostcodeAgg sheet …")
    build_pq_agg_sheet(wb)

    print("Building Tool sheet …")
    build_tool_sheet(wb)

    # Ensure Tool sheet is the active sheet on open
    for i, sh in enumerate(wb.worksheets):
        sh.sheet_view.tabSelected = (sh.title == "Tool")
    wb.active = wb["Tool"]

    print(f"Saving to {OUT} …")
    wb.save(OUT)

    print("Injecting Power Query …")
    inject_power_query(OUT)

    print(f"\nDone!  →  {OUT}")


if __name__ == "__main__":
    main()
